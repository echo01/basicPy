from flask_cors import CORS
from flask import Flask, make_response, request
import json
import os
from bson.objectid import ObjectId
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
import mongo
import gen_excel



app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
template_loc = os.environ.get('ROOT_APP', '/Git/Pimus_webscada')


@app.route("/")
def hello_world():
    return "Hello, World!"


@app.route("/export/<template_id>", methods=['POST'])
def export_excel(template_id):
    request_data = json.loads(request.get_data())
    client = mongo.get_database()
    collections = client.get_collection('upload_record_templates')
    data = collections.find_one({'_id': ObjectId(template_id)})
    file_loc = '{}/{}'.format(template_loc, data['location'])
    wb = load_workbook(file_loc)
    ws = wb[wb.sheetnames[0]]
    tagList = []
    for x in request_data['tagList']:
        tagList.append(x['device'])
    record = gen_excel.prepareData(request_data['data'], tagList)
    gen_excel.fill_data(ws, record)
    response = make_response(save_virtual_workbook(wb))
    response.headers['Content-Type'] = 'application/ms-excel'
    response.headers['Content-Disposition'] = 'attachment; filename={}'.format('report.xlsx')
    return response

if __name__ == "__main__":
    app.run(debug=True)